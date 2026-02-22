import streamlit as st
import pandas as pd
import io
import zipfile
import unicodedata
import re
import isodate 
import datetime
from googleapiclient.discovery import build 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# --- ヘルパー関数 ---

def from_hex_fill(hex_code):
    return PatternFill(start_color=hex_code, end_color=hex_code, fill_type="solid")

def get_display_width(text):
    if not isinstance(text, str):
        text = str(text)
    width = 0
    for char in text:
        if unicodedata.east_asian_width(char) in ('F', 'W', 'A'):
            width += 2
        else:
            width += 1
    return width

def extract_video_id(url):
    """YouTubeのURLから動画IDを抽出する"""
    if not isinstance(url, str):
        return None
    patterns = [
        r'(?:v=|\/)([0-9A-Za-z_-]{11}).*',
        r'(?:youtu\.be\/)([0-9A-Za-z_-]{11})',
        r'(?:embed\/)([0-9A-Za-z_-]{11})'
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None

def extract_all_urls(cell_value):
    """セル内のテキストからすべてのURLを抽出する"""
    if not isinstance(cell_value, str) or cell_value.strip().lower() == "nan":
        return []
    # httpまたはhttpsで始まるURLをすべて抽出（区切り文字に依存しない）
    urls = re.findall(r'https?://[^\s,、\u3000]+', cell_value)
    return urls

def fetch_youtube_details(api_key, video_ids):
    """YouTube Data APIを使用して動画の詳細を一括取得する"""
    if not api_key or not video_ids:
        return {}
    
    youtube = build('youtube', 'v3', developerKey=api_key)
    results = {}
    
    chunk_size = 50
    for i in range(0, len(video_ids), chunk_size):
        chunk = video_ids[i:i+chunk_size]
        try:
            request = youtube.videos().list(
                part="contentDetails,status",
                id=",".join(chunk)
            )
            response = request.execute()
            
            for item in response.get("items", []):
                vid = item["id"]
                duration_iso = item["contentDetails"]["duration"]
                privacy_status = item["status"]["privacyStatus"]
                results[vid] = {
                    "duration": duration_iso,
                    "status": privacy_status
                }
        except Exception as e:
            st.error(f"YouTube API通信エラー: {e}")
            
    return results

def format_duration(iso_duration):
    """ISO 8601形式を変換"""
    try:
        dur = isodate.parse_duration(iso_duration)
        total_seconds = int(dur.total_seconds())
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes}分{seconds}秒"
    except:
        return ""

def format_total_seconds(total_seconds):
    """合計秒数を「○分○秒」形式に変換する"""
    if total_seconds <= 0:
        return ""
    minutes = total_seconds // 60
    seconds = total_seconds % 60
    return f"{minutes}分{seconds}秒"

def _to_circled_num(n):
    """数値を丸付き数字（①〜⑳）に変換する"""
    if 1 <= n <= 20:
        return chr(0x245f + n)
    return str(n)

def _from_circled_num(s):
    """丸付き数字（①〜⑳）から数値に変換する"""
    if len(s) == 1 and 0x2460 <= ord(s) <= 0x2473:
        return ord(s) - 0x245f
    try:
        return int(s)
    except:
        return 1

# --- メインアプリ ---

st.set_page_config(page_title="録画審査表ジェネレーター", layout="wide")

st.title("🗂️ 録画審査表ジェネレーター")
st.markdown("""
アップロードされた名簿Excelファイルから、部門ごとの採点用ファイルを生成します。
**特徴:**
- YouTube API連携により、動画時間と再生可否（公開設定）を自動チェックします。
- 講評欄の文字数設定に応じてヘッダーが自動で変わります。
- 処理結果ログを含むZIPファイルを生成します。
""")

# --- APIキー設定 ---
with st.expander("🔑 YouTube API設定 (必須)", expanded=True):
    secret_key = st.secrets.get("YOUTUBE_API_KEY", None)
    user_input_key = st.text_input(
        "YouTube Data APIキー（Secrets設定済みの場合は空欄でOKです）", 
        type="password", 
        help="Google Cloud Consoleで取得したAPIキーを入力してください。"
    )
    final_api_key = user_input_key if user_input_key else secret_key
    
    if user_input_key:
        st.info("ℹ️ 入力されたAPIキーを使用します")
    elif secret_key:
        st.success("✅ Secrets設定済みのAPIキーが適用されています")
    else:
        st.warning("⚠️ APIキーが設定されていません。動画情報の自動取得機能は動作しません。")

# --- 1. ファイルアップロード ---
uploaded_file = st.file_uploader("出場者名簿（Excelファイル）をアップロードしてください", type=["xlsx"])

if uploaded_file:
    try:
        xls = pd.ExcelFile(uploaded_file)
        all_sheets = xls.sheet_names

        st.divider()
        st.subheader("1. 対象シートの選択")
        
        ignore_keywords = ["原本", "総合名簿", "削除ログ", "ログ"]
        default_selections = [s for s in all_sheets if not any(kw in s for kw in ignore_keywords)]
        
        target_sheets = st.multiselect(
            "審査表を作成したいシート（部門）を選択してください",
            options=all_sheets,
            default=default_selections
        )

        if target_sheets:
            df_sample = pd.read_excel(xls, sheet_name=target_sheets[0])
            source_columns = ["（なし）"] + list(df_sample.columns)

            st.divider()
            st.subheader("2. 列のマッピングと出力設定")

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("##### 📋 列の紐付け")
                
                def get_index(options, keywords):
                    for i, opt in enumerate(options):
                        for kw in keywords:
                            if kw in opt:
                                return i
                    return 0

                mapping = {}
                mapping["entry_number"] = st.selectbox("出場番号", source_columns, index=get_index(source_columns, ["番号", "No", "ID"]))
                mapping["entry_name"] = st.selectbox("出場者名", source_columns, index=get_index(source_columns, ["氏名", "名前", "団体名"]))
                mapping["instrument"] = st.selectbox("楽器名 (任意)", source_columns, index=get_index(source_columns, ["楽器"]))
                mapping["age"] = st.selectbox("年齢", source_columns, index=get_index(source_columns, ["年齢", "学年"]))
                mapping["song"] = st.selectbox("曲目", source_columns, index=get_index(source_columns, ["曲目", "曲名"]))
                mapping["youtube"] = st.selectbox("YouTube URL", source_columns, index=get_index(source_columns, ["YouTube", "URL", "動画"]))
                mapping["duration"] = st.selectbox("演奏時間 (元データがあれば)", source_columns, index=get_index(source_columns, ["時間", "タイム"]))
                # 【追加】メールアドレス列の指定
                mapping["email"] = st.selectbox("メールアドレス (任意・連絡用)", source_columns, index=get_index(source_columns, ["メール", "mail", "Email"]))

            with col2:
                st.markdown("##### ⚙️ 審査表の出力設定")
                
                output_filename_base = st.text_input("出力ファイル名の基本名", value="録画審査表")
                score_mode = st.selectbox("採点方式", ["採点(100点満点)", "採点(◯△✕)"])
                score_header_display = "採点"
                
                min_char_count = st.number_input("講評の最低文字数（警告用）", min_value=0, value=100, step=10)
                
                if min_char_count > 0:
                    comment_header_text = f"審査講評（{min_char_count}文字以上）"
                else:
                    comment_header_text = "審査講評（100～200文字程度以上）"
                
                st.info(f"出力されるヘッダー名: **{comment_header_text}**")

            # --- 実行ボタン ---
            st.divider()
            generate_btn = st.button("審査表を作成する", type="primary")

            if generate_btn:
                if any(mapping[k] == "（なし）" for k in ["entry_number", "entry_name", "song", "youtube"]):
                    st.error("エラー: 必須項目（番号、氏名、曲目、URL）には列を指定してください。")
                elif not final_api_key:
                     st.error("エラー: YouTube APIキーが設定されていません。")
                else:
                    output_files = {}
                    error_logs_list = [] # 構造化されたログデータ用
                    progress_bar = st.progress(0)
                    
                    try:
                        total_sheets = len(target_sheets)
                        
                        for i, sheet_name in enumerate(target_sheets):
                            df = pd.read_excel(xls, sheet_name=sheet_name)
                            
                            missing_cols = []
                            for k, v in mapping.items():
                                if v != "（なし）" and v not in df.columns:
                                    missing_cols.append(v)
                            
                            if missing_cols:
                                st.warning(f"シート「{sheet_name}」には以下の列が存在しないためスキップしました: {', '.join(missing_cols)}")
                                continue

                            # YouTube API処理（複数URL対応）
                            id_map = {}  # idx -> [(vid, url), ...]
                            if mapping["youtube"] != "（なし）":
                                for idx, row in df.iterrows():
                                    cell_value = row[mapping["youtube"]]
                                    urls = extract_all_urls(str(cell_value) if pd.notna(cell_value) else "")
                                    vid_url_pairs = []
                                    for url in urls:
                                        vid = extract_video_id(url)
                                        if vid:
                                            vid_url_pairs.append((vid, url))
                                        else:
                                            vid_url_pairs.append((None, url))  # ID抽出不可
                                    if vid_url_pairs:
                                        id_map[idx] = vid_url_pairs
                            
                            # 全動画IDを一括取得
                            all_vids = []
                            for pairs in id_map.values():
                                for vid, url in pairs:
                                    if vid:
                                        all_vids.append(vid)
                            unique_ids = list(set(all_vids))
                            api_results = fetch_youtube_details(final_api_key, unique_ids)
                            
                            # シート内の最大リンク数を事前計算
                            max_links = 0
                            for idx, row in df.iterrows():
                                if idx in id_map:
                                    max_links = max(max_links, len(id_map[idx]))
                                else:
                                    cell_value = row[mapping["youtube"]] if mapping["youtube"] != "（なし）" else ""
                                    urls = extract_all_urls(str(cell_value) if pd.notna(cell_value) else "")
                                    max_links = max(max_links, len(urls))
                            if max_links == 0:
                                max_links = 1  # 最低1列は確保
                            
                            new_data = []
                            for idx, row in df.iterrows():
                                num_val = row[mapping["entry_number"]] if mapping["entry_number"] != "（なし）" else ""
                                name_val = row[mapping["entry_name"]] if mapping["entry_name"] != "（なし）" else ""
                                cell_value = row[mapping["youtube"]] if mapping["youtube"] != "（なし）" else ""
                                email_val = row[mapping["email"]] if mapping["email"] != "（なし）" else "不明"
                                
                                duration_text = ""
                                if mapping["duration"] != "（なし）":
                                    duration_text = row[mapping["duration"]]

                                # 複数URL対応: 各リンクの情報を収集
                                video_links = []  # [(link_text, url), ...]
                                total_seconds = 0
                                has_api_duration = False
                                has_error = False
                                
                                if idx in id_map:
                                    pairs = id_map[idx]
                                    for vid, url in pairs:
                                        if vid and vid in api_results:
                                            details = api_results[vid]
                                            status = details["status"]
                                            if status in ['public', 'unlisted']:
                                                try:
                                                    dur = isodate.parse_duration(details["duration"])
                                                    total_seconds += int(dur.total_seconds())
                                                    has_api_duration = True
                                                except:
                                                    pass
                                                video_links.append((url, True))  # 有効
                                            else:
                                                error_msg = f"動画設定が「{status}」のため再生できません"
                                                error_logs_list.append({
                                                    "type": "error",
                                                    "dept": sheet_name,
                                                    "no": num_val,
                                                    "name": name_val,
                                                    "reason": error_msg,
                                                    "url": url,
                                                    "email": email_val
                                                })
                                                video_links.append((url, False))  # 無効
                                                has_error = True
                                        elif vid and vid not in api_results:
                                            error_msg = "動画が見つかりません（削除またはID無効）"
                                            error_logs_list.append({
                                                "type": "error",
                                                "dept": sheet_name,
                                                "no": num_val,
                                                "name": name_val,
                                                "reason": error_msg,
                                                "url": url,
                                                "email": email_val
                                            })
                                            video_links.append((url, False))  # 無効
                                            has_error = True
                                        else:
                                            # vid is None: YouTube以外のURLまたは形式不明
                                            error_msg = "URLの形式が不明です"
                                            error_logs_list.append({
                                                "type": "error",
                                                "dept": sheet_name,
                                                "no": num_val,
                                                "name": name_val,
                                                "reason": error_msg,
                                                "url": url,
                                                "email": email_val
                                            })
                                            video_links.append((url, False))  # 無効
                                            has_error = True
                                else:
                                    # id_mapにない場合: URLがないか抽出できなかった
                                    raw_urls = extract_all_urls(str(cell_value) if pd.notna(cell_value) else "")
                                    if raw_urls:
                                        for url in raw_urls:
                                            error_msg = "URLの形式が不明です"
                                            error_logs_list.append({
                                                "type": "error",
                                                "dept": sheet_name,
                                                "no": num_val,
                                                "name": name_val,
                                                "reason": error_msg,
                                                "url": url,
                                                "email": email_val
                                            })
                                
                                # 演奏時間: API結果があれば合算値を使用
                                if has_api_duration:
                                    duration_text = format_total_seconds(total_seconds)
                                    if has_error:
                                        duration_text += "（一部要確認）"
                                elif has_error:
                                    duration_text = "【要確認】"
                                
                                # DataFrame構築
                                record = {
                                    "出場部門": sheet_name,
                                    "出場番号": num_val,
                                    "出場者名": name_val,
                                    "年齢": row[mapping["age"]] if mapping["age"] != "（なし）" else "",
                                    "曲目": row[mapping["song"]] if mapping["song"] != "（なし）" else "",
                                }
                                if mapping["instrument"] != "（なし）":
                                    record["楽器名"] = row[mapping["instrument"]]
                                
                                # 動画列の動的生成
                                link_count = len(video_links)
                                for n in range(max_links):
                                    if max_links == 1:
                                        col_video = "動画"
                                        col_url = "YouTube URL"
                                    else:
                                        col_video = f"動画{_to_circled_num(n + 1)}"
                                        col_url = f"YouTube URL {n + 1}"
                                    
                                    if n < link_count:
                                        url, is_valid = video_links[n]
                                        if max_links == 1:
                                            record[col_video] = "再生"
                                        else:
                                            record[col_video] = f"再生{_to_circled_num(n + 1)}"
                                        record[col_url] = url
                                    else:
                                        record[col_video] = ""
                                        record[col_url] = ""
                                
                                record["演奏時間"] = duration_text
                                record[score_header_display] = ""
                                record[comment_header_text] = ""
                                
                                new_data.append(record)
                            
                            df_out = pd.DataFrame(new_data)
                            
                            # 列順序の動的構築
                            cols_order = ["出場部門"]
                            if mapping["instrument"] != "（なし）":
                                cols_order.append("楽器名")
                            cols_order.extend(["出場番号", "出場者名", "年齢", "曲目"])
                            # 動画列とURL列を交互に配置
                            for n in range(max_links):
                                if max_links == 1:
                                    cols_order.append("動画")
                                    cols_order.append("YouTube URL")
                                else:
                                    cols_order.append(f"動画{_to_circled_num(n + 1)}")
                                    cols_order.append(f"YouTube URL {n + 1}")
                            cols_order.extend(["演奏時間", score_header_display, comment_header_text])
                            
                            final_cols = [c for c in cols_order if c in df_out.columns]
                            df_out = df_out[final_cols]

                            wb = Workbook()
                            ws = wb.active
                            ws.title = "審査表"

                            for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), 1):
                                # 行の高さ自動調整
                                if r_idx > 1:
                                    max_lines = 1
                                    for val in row:
                                        val_str = str(val) if val is not None else ""
                                        lines = val_str.count('\n') + 1
                                        if lines > max_lines:
                                            max_lines = lines
                                    row_height = max(30, max_lines * 15)
                                    ws.row_dimensions[r_idx].height = row_height

                                for c_idx, value in enumerate(row, 1):
                                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                    col_name = df_out.columns[c_idx - 1]
                                    
                                    thin = Side(border_style="thin", color="000000")
                                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                                    if r_idx == 1: 
                                        cell.font = Font(bold=True, color="FFFFFF")
                                        cell.fill = from_hex_fill("4F81BD")
                                        cell.alignment = Alignment(horizontal="left", vertical="center")
                                    else: 
                                        align_h = "center" if col_name == "年齢" or col_name.startswith("動画") or col_name == score_header_display else "left"
                                        cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=True)
                                        
                                        # 【複数対応】「動画」「動画①」等のハイパーリンク設定
                                        if col_name.startswith("動画") and value and str(value).startswith("再生"):
                                            # 対応するURL列名を特定
                                            if col_name == "動画":
                                                url_col_name = "YouTube URL"
                                            else:
                                                # 「動画①」→「YouTube URL 1」等
                                                num_suffix = col_name.replace("動画", "")
                                                url_num = _from_circled_num(num_suffix)
                                                url_col_name = f"YouTube URL {url_num}"
                                            if url_col_name in df_out.columns:
                                                url_val = df_out.iloc[r_idx-2][url_col_name]
                                                if url_val and str(url_val).lower() != "nan" and str(url_val).strip():
                                                    cell.hyperlink = str(url_val)
                                                    cell.font = Font(color="0563C1", underline="single")
                                        
                                        # 演奏時間のエラー強調
                                        if col_name == "演奏時間" and ("【" in str(value) or "確認" in str(value)):
                                            cell.font = Font(color="FF0000", bold=True)

                            # 列幅と非表示設定
                            for i_col, col_name in enumerate(final_cols):
                                column_letter = ws.cell(row=1, column=i_col+1).column_letter
                                
                                # 【変更】YouTube URL列は非表示にする
                                if col_name.startswith("YouTube URL"):
                                    ws.column_dimensions[column_letter].hidden = True
                                    continue # 幅設定不要
                                
                                if col_name == "出場番号":
                                    ws.column_dimensions[column_letter].width = 12
                                elif col_name == "年齢":
                                    ws.column_dimensions[column_letter].width = 8
                                elif col_name.startswith("動画"): # 動画列（動的拡張対応）
                                    ws.column_dimensions[column_letter].width = 10
                                elif col_name == comment_header_text:
                                    ws.column_dimensions[column_letter].width = 50
                                elif col_name == score_header_display:
                                    ws.column_dimensions[column_letter].width = 10
                                else:
                                    # 【変更】余白計算ロジック: 最大文字数 + 2 (全角1文字分)
                                    data_lengths = [get_display_width(str(val)) for val in df_out[col_name].fillna("")]
                                    if data_lengths:
                                        max_len = max(data_lengths)
                                        # 固定加算方式に変更
                                        calc_width = max_len + 3 
                                        limit_width = 80
                                        final_width = max(min(calc_width, limit_width), 10)
                                        ws.column_dimensions[column_letter].width = final_width
                                    else:
                                        ws.column_dimensions[column_letter].width = 20

                            # 入力規則
                            comment_col_idx = None
                            for cell in ws[1]:
                                if cell.value == comment_header_text:
                                    comment_col_idx = cell.column_letter
                                    break
                            
                            if min_char_count > 0 and comment_col_idx:
                                formula = f'LEN({comment_col_idx}2)>={min_char_count}'
                                dv = DataValidation(
                                    type="custom",
                                    formula1=formula,
                                    allow_blank=True,
                                    showErrorMessage=True,
                                    errorTitle="入力文字数不足",
                                    error="審査講評は指定された文字数以上入力してください。"
                                )
                                dv.add(f"{comment_col_idx}2:{comment_col_idx}{len(df_out)+1}")
                                ws.add_data_validation(dv)

                            excel_buffer = io.BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            output_files[f"{output_filename_base}_{sheet_name}.xlsx"] = excel_buffer
                            progress_val = min((i + 1) / total_sheets, 1.0)
                            progress_bar.progress(progress_val)

                        # --- ログファイルの生成 (体裁を整える) ---
                        
                        log_lines = []
                        log_lines.append("【再生可否判定レポート】")
                        log_lines.append(f"確認日時: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M')}")
                        log_lines.append("\n" + "-"*50)
                        log_lines.append("⚠️ 要確認（再生不可など）")
                        log_lines.append("-"*50 + "\n")
                        
                        if error_logs_list:
                            for log in error_logs_list:
                                log_lines.append(f"[{log['dept']}] {log['no']} {log['name']} 様")
                                log_lines.append(f"状況: {log['reason']}")
                                log_lines.append(f"URL : {log['url']}")
                                log_lines.append(f"Email: {log['email']}")
                                log_lines.append("") # 空行
                        else:
                            log_lines.append("（該当なし。すべての動画が正常に確認されました）\n")
                            
                        log_lines.append("\n" + "-"*50)
                        log_lines.append("✅ 確認完了")
                        log_lines.append("-"*50)
                        log_lines.append("上記以外の動画については、正常に時間取得が完了しています。")

                        log_content = "\n".join(log_lines)
                        
                        # ファイル名変更: 再生可否判定.txt
                        log_buffer = io.BytesIO()
                        log_buffer.write(log_content.encode('utf-8-sig'))
                        log_buffer.seek(0)
                        output_files["再生可否判定.txt"] = log_buffer

                        st.success("作成が完了しました！ZIPファイルをダウンロードしてください。")
                        
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, "w") as zf:
                            for fname, fbuff in output_files.items():
                                zf.writestr(fname, fbuff.getvalue())
                        zip_buffer.seek(0)
                        
                        st.download_button(
                            label="📥 審査表セットをダウンロード (ZIP)",
                            data=zip_buffer,
                            file_name=f"{output_filename_base}_セット.zip",
                            mime="application/zip"
                        )
                        
                        if error_logs_list:
                            st.error(f"⚠️ {len(error_logs_list)}件の動画に問題が見つかりました。詳細は「再生可否判定.txt」をご確認ください。")
                            # 簡易表示
                            simple_log = "\n".join([f"[{l['dept']}] {l['name']}: {l['reason']}" for l in error_logs_list])
                            st.text_area("エラー詳細ログ（プレビュー）", value=simple_log, height=150)

                    except Exception as e:
                        st.error(f"処理中にエラーが発生しました: {e}")

    except Exception as e:
        st.error(f"ファイルの読み込みに失敗しました: {e}")
